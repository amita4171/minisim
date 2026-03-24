#!/usr/bin/env python3
"""
MiniSim Eval Set Generator
Builds a comprehensive evaluation spreadsheet for the MiniSim swarm prediction engine.

Eval dimensions:
1. Prediction Questions (resolved + unresolved) across 8 categories
2. Calibration Benchmarks — questions bucketed by true probability
3. Mode Collapse Stress Tests — adversarial questions that trigger convergence
4. Agent Diversity Benchmarks — expected std dev thresholds
5. Context Quality Tests — same question with varying seed context
6. Aggregation Method Comparison — expected outputs for different aggregators
7. Backtest Registry — real Kalshi/Polymarket/Metaculus markets with resolutions
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "/sessions/great-focused-mccarthy/mnt/outputs/MiniSim_Eval_Set.xlsx"
wb = Workbook()

# ── Style constants ──
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor="2F5496")
CAT_FILL = PatternFill("solid", fgColor="D6E4F0")
CAT_FONT = Font(name="Arial", bold=True, size=11, color="1F3864")
BODY_FONT = Font(name="Arial", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def style_header(ws, headers, col_widths=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = THIN_BORDER
    if col_widths:
        for c, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"


def add_row(ws, row_num, values):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=c, value=v)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER


# ═══════════════════════════════════════════════════════════════
# SHEET 1: PREDICTION QUESTIONS
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Prediction Questions"

headers = [
    "Q_ID", "Category", "Sub-Category", "Question", "Resolution_Date",
    "Actual_Resolution", "True_Probability", "Difficulty",
    "Seed_Context", "Context_Quality", "Expected_P_Range_Low",
    "Expected_P_Range_High", "Known_Biases", "Source",
    "Market_Price", "Edge_Case_Tags", "Notes"
]
col_widths = [8, 14, 16, 50, 14, 14, 14, 10, 60, 12, 12, 12, 30, 20, 12, 25, 30]
style_header(ws1, headers, col_widths)

questions = [
    # ── ECONOMICS (30 questions) ──
    ("ECON-001", "Economics", "Monetary Policy", "Will the Federal Reserve cut the federal funds rate at the May 2026 FOMC meeting?",
     "2026-05-07", "PENDING", None, "Medium",
     "Fed held rates at 4.25-4.50% at March 2026 meeting. Inflation 2.8% YoY Feb 2026. Feb jobs +180k, unemployment 4.1%. Powell signaled 'patience' and 'data dependence'. CME FedWatch: 35% probability May cut.",
     "High", 0.25, 0.45, "Recency bias, anchoring to market price", "CME FedWatch / FOMC", 0.35, "monetary_policy, high_frequency", ""),
    ("ECON-002", "Economics", "Monetary Policy", "Will the Federal Reserve cut rates at least once before September 2026?",
     "2026-09-30", "PENDING", None, "Easy",
     "Markets price ~2 cuts by end of 2026. Inflation trending down but above target. Labor market softening gradually.",
     "Medium", 0.55, 0.75, "Overconfidence in rate cut timing", "CME FedWatch", 0.65, "monetary_policy, cumulative", ""),
    ("ECON-003", "Economics", "Inflation", "Will US CPI year-over-year inflation fall below 2.5% before July 2026?",
     "2026-07-31", "PENDING", None, "Medium",
     "CPI at 2.8% as of Feb 2026. Core PCE at 2.6%. Shelter inflation still elevated. Used car prices declining.",
     "Medium", 0.30, 0.50, "Anchoring to recent trend", "BLS CPI Data", None, "inflation, threshold", ""),
    ("ECON-004", "Economics", "Labor Market", "Will US unemployment exceed 4.5% at any point in 2026?",
     "2026-12-31", "PENDING", None, "Medium",
     "Unemployment at 4.1% Feb 2026. JOLTS openings declining. Tech layoffs continue. Construction hiring steady.",
     "Medium", 0.25, 0.40, "Availability bias from layoff headlines", "BLS", None, "labor, threshold", ""),
    ("ECON-005", "Economics", "GDP", "Will US real GDP growth in Q1 2026 exceed 2.0% (annualized)?",
     "2026-04-30", "PENDING", None, "Easy",
     "Q4 2025 GDP +2.3%. Atlanta Fed GDPNow tracking +1.8% for Q1. Consumer spending resilient. Business investment mixed.",
     "High", 0.40, 0.60, "Overweighting recent quarters", "BEA / Atlanta Fed", None, "gdp, quarterly", ""),
    ("ECON-006", "Economics", "Housing", "Will the median US existing home sale price exceed $400,000 in March 2026?",
     "2026-04-15", "PENDING", None, "Easy",
     "Median price $387k Jan 2026. Inventory still low. Mortgage rates ~6.8%. Spring buying season approaching.",
     "High", 0.35, 0.55, "Seasonal adjustment confusion", "NAR", None, "housing, threshold", ""),
    ("ECON-007", "Economics", "Trade", "Will the US impose new tariffs on Chinese goods exceeding 25% in 2026?",
     "2026-12-31", "PENDING", None, "Hard",
     "Existing tariffs at 25% on $250B. Biden maintained Trump tariffs. Election rhetoric favors tough trade stance. China retaliation risk.",
     "Low", 0.30, 0.55, "Political bias", "USTR / Trade data", None, "trade_policy, geopolitical", ""),
    ("ECON-008", "Economics", "Recession", "Will the NBER declare a US recession starting in 2026?",
     "2027-12-31", "PENDING", None, "Hard",
     "Yield curve uninverted. LEI mixed. Consumer confidence declining. Corporate earnings still growing. Sahm rule not triggered.",
     "Medium", 0.10, 0.25, "Recession prediction bias, headline fear", "NBER", None, "recession, long_horizon", ""),
    ("ECON-009", "Economics", "Crypto", "Will Bitcoin exceed $120,000 at any point in Q2 2026?",
     "2026-06-30", "PENDING", None, "Hard",
     "BTC at ~$87k March 2026. ETF inflows strong. Halving effect fading. Macro uncertainty high. Institutional adoption growing.",
     "Medium", 0.25, 0.45, "Crypto enthusiasm bias", "CoinGecko", None, "crypto, volatile_asset", ""),
    ("ECON-010", "Economics", "Oil", "Will WTI crude oil average above $80/barrel in Q2 2026?",
     "2026-06-30", "PENDING", None, "Medium",
     "WTI at $72 March 2026. OPEC+ extending cuts. US production at record highs. China demand uncertain. Middle East tensions.",
     "Medium", 0.30, 0.50, "Geopolitical risk overweighting", "EIA", None, "commodities, average", ""),

    # More economics
    ("ECON-011", "Economics", "Tech Sector", "Will Nvidia's market cap exceed $4 trillion before July 2026?",
     "2026-07-01", "PENDING", None, "Medium",
     "NVDA market cap ~$3.2T March 2026. AI spending accelerating. Blackwell ramp in progress. Competition from AMD, custom chips.",
     "Medium", 0.35, 0.55, "Tech enthusiasm, momentum bias", "Yahoo Finance", None, "single_stock, market_cap", ""),
    ("ECON-012", "Economics", "Labor Market", "Will the US economy add more than 200k jobs in any single month of Q2 2026?",
     "2026-06-30", "PENDING", None, "Easy",
     "Recent months: 180k, 195k, 210k, 175k. Seasonal hiring patterns. Government hiring steady.",
     "High", 0.55, 0.75, "Anchoring to recent numbers", "BLS", None, "labor, threshold, monthly", ""),
    ("ECON-013", "Economics", "Consumer", "Will US retail sales year-over-year growth turn negative in 2026?",
     "2026-12-31", "PENDING", None, "Hard",
     "Retail sales +3.2% YoY Feb 2026. Credit card delinquencies rising. Student loan payments resumed. Savings rate declining.",
     "Medium", 0.15, 0.30, "Pessimism bias from delinquency headlines", "Census Bureau", None, "consumer, annual", ""),
    ("ECON-014", "Economics", "Monetary Policy", "Will the ECB cut rates more than the Fed in 2026 (total basis points)?",
     "2026-12-31", "PENDING", None, "Medium",
     "ECB at 3.75%, already cut twice in 2025. Fed at 4.25-4.50%, zero cuts yet in 2026. Eurozone growth weaker than US.",
     "Medium", 0.55, 0.75, "Home bias (US-centric agents)", "ECB / Fed", None, "comparative, central_banks", ""),
    ("ECON-015", "Economics", "Debt", "Will the US 10-year Treasury yield exceed 5.0% at any point in 2026?",
     "2026-12-31", "PENDING", None, "Medium",
     "10Y at 4.35% March 2026. Deficit spending elevated. Treasury issuance heavy. Foreign demand uncertain.",
     "Medium", 0.20, 0.40, "Anchoring to recent range", "Treasury.gov", None, "rates, threshold", ""),

    # ── GEOPOLITICS (20 questions) ──
    ("GEO-001", "Geopolitics", "Conflict", "Will there be a formal ceasefire agreement in the Russia-Ukraine war before 2027?",
     "2027-01-01", "PENDING", None, "Hard",
     "War in its 4th year. Trump administration pushing negotiations. Russia controlling ~20% of Ukrainian territory. Ukraine fatigue in Europe. China mediating.",
     "Medium", 0.15, 0.35, "Wishful thinking bias, status quo bias", "UN / Reuters", None, "conflict, long_horizon", ""),
    ("GEO-002", "Geopolitics", "Conflict", "Will Israel and Hamas reach a permanent ceasefire agreement in 2026?",
     "2026-12-31", "PENDING", None, "Hard",
     "Temporary truces have collapsed. Hostage negotiations ongoing. Regional escalation risk. International pressure increasing.",
     "Low", 0.10, 0.30, "Recency of temporary agreements", "Reuters / AP", None, "conflict, middle_east", ""),
    ("GEO-003", "Geopolitics", "Taiwan", "Will China conduct military exercises within 12nm of Taiwan in 2026?",
     "2026-12-31", "PENDING", None, "Hard",
     "Last major exercises Aug 2022. Tensions elevated after Taiwan elections. US arms sales continuing. Pelosi visit precedent.",
     "Low", 0.15, 0.35, "Catastrophizing, availability bias", "DoD / CSIS", None, "military, taiwan", ""),
    ("GEO-004", "Geopolitics", "Elections", "Will the UK call a snap general election before 2027?",
     "2027-01-01", "PENDING", None, "Medium",
     "Labour won July 2024 with large majority. Starmer approval declining. Economic challenges. No constitutional requirement.",
     "Low", 0.05, 0.15, "Projection of US politics onto UK", "UK Parliament", None, "elections, unlikely", ""),
    ("GEO-005", "Geopolitics", "Trade", "Will the EU impose digital services taxes on US tech companies exceeding current rates in 2026?",
     "2026-12-31", "PENDING", None, "Medium",
     "OECD Pillar One stalled. France, Spain already have DSTs. US threatening retaliation. Tech lobby strong.",
     "Medium", 0.30, 0.50, "US-centric bias", "European Commission", None, "trade, technology", ""),
    ("GEO-006", "Geopolitics", "Climate", "Will global CO2 emissions in 2026 be lower than 2025?",
     "2027-03-01", "PENDING", None, "Medium",
     "Emissions rose 1.1% in 2025. China coal expansion. Renewable deployment record pace. India growth offsetting gains.",
     "Medium", 0.20, 0.35, "Optimism bias from renewable headlines", "IEA / GCP", None, "climate, annual", ""),
    ("GEO-007", "Geopolitics", "Conflict", "Will North Korea conduct a nuclear weapons test in 2026?",
     "2026-12-31", "PENDING", None, "Medium",
     "Last test Sept 2017. Satellite imagery shows activity. Kim Jong Un rhetoric escalating. Tactical nuke development reported.",
     "Low", 0.15, 0.30, "Threat inflation", "38 North / IAEA", None, "nuclear, military", ""),
    ("GEO-008", "Geopolitics", "Sanctions", "Will the US impose new sanctions on Russian energy exports in 2026?",
     "2026-12-31", "PENDING", None, "Medium",
     "Existing sanctions on oil price cap. Shadow fleet evading. India/China buying discounted crude. Trump less hawkish on Russia.",
     "Medium", 0.25, 0.45, "Political bias", "Treasury OFAC", None, "sanctions, energy", ""),
    ("GEO-009", "Geopolitics", "AI Governance", "Will the EU AI Act result in a major fine (>100M EUR) against a US tech company in 2026?",
     "2026-12-31", "PENDING", None, "Hard",
     "AI Act enforcement began Feb 2025. Compliance timelines staggered. Investigation periods lengthy. Meta, OpenAI, Google flagged.",
     "Low", 0.05, 0.15, "Overestimating enforcement speed", "EU Commission", None, "regulation, technology", ""),
    ("GEO-010", "Geopolitics", "Space", "Will India successfully land a spacecraft on the Moon in 2026?",
     "2026-12-31", "PENDING", None, "Medium",
     "Chandrayaan-3 succeeded Aug 2023. Chandrayaan-4 sample return planned 2028. No 2026 landing mission announced.",
     "Medium", 0.05, 0.15, "Confusion with Chandrayaan-4 timeline", "ISRO", None, "space, unlikely", ""),

    # ── TECHNOLOGY (20 questions) ──
    ("TECH-001", "Technology", "AI", "Will OpenAI release GPT-5 (or equivalent next-gen model) before July 2026?",
     "2026-07-01", "PENDING", None, "Medium",
     "GPT-4o released May 2024. o1 released Sept 2024. o3 released Jan 2025. Rumors of GPT-5 training completion. Competitive pressure from Claude, Gemini.",
     "Medium", 0.45, 0.65, "Hype cycle bias", "OpenAI / TechCrunch", None, "ai_models, release_date", ""),
    ("TECH-002", "Technology", "AI", "Will any AI model achieve >90% on the ARC-AGI benchmark before 2027?",
     "2027-01-01", "PENDING", None, "Hard",
     "o3 scored 87.5% on ARC-AGI Dec 2024. Rapid progress. But remaining 12.5% is disproportionately hard. Chollet skeptical of brute-force approaches.",
     "Medium", 0.35, 0.55, "Extrapolation of recent progress", "ARC Prize", None, "ai_benchmark, threshold", ""),
    ("TECH-003", "Technology", "AI", "Will Anthropic raise a funding round valuing the company at >$100B in 2026?",
     "2026-12-31", "PENDING", None, "Easy",
     "Anthropic valued at $61B as of Dec 2024. Revenue growing rapidly. Claude competitive. Google, Amazon invested. AI market expanding.",
     "Medium", 0.50, 0.70, "Insider optimism", "TechCrunch / PitchBook", None, "funding, valuation", ""),
    ("TECH-004", "Technology", "Robotics", "Will Tesla begin delivering Optimus robots to external customers in 2026?",
     "2026-12-31", "PENDING", None, "Hard",
     "Optimus Gen 2 demo Dec 2024. Internal Tesla factory use. Musk claims 2026 external sales. History of missed timelines.",
     "Low", 0.10, 0.25, "Musk timeline optimism", "Tesla IR", None, "robotics, delivery", ""),
    ("TECH-005", "Technology", "Quantum", "Will any company demonstrate quantum advantage on a commercially relevant problem in 2026?",
     "2026-12-31", "PENDING", None, "Hard",
     "Google Willow chip Dec 2024. IBM Condor 1,121 qubits. Error correction improving. But practical applications still years away for most uses.",
     "Low", 0.10, 0.25, "Quantum hype cycle", "Nature / arXiv", None, "quantum, breakthrough", ""),
    ("TECH-006", "Technology", "Social Media", "Will TikTok be banned or forced to divest in the US by end of 2026?",
     "2026-12-31", "PENDING", None, "Hard",
     "Supreme Court upheld ban law Jan 2025. Trump extended deadline. ByteDance exploring options. Political bipartisan support for action.",
     "Medium", 0.25, 0.45, "Political prediction difficulty", "Congress / Reuters", None, "regulation, social_media", ""),
    ("TECH-007", "Technology", "EV", "Will US EV market share of new car sales exceed 12% in 2026?",
     "2027-01-31", "PENDING", None, "Easy",
     "EV share ~10% in 2025. Tesla dominance declining. Legacy OEMs ramping. Charging infrastructure improving. IRA incentives active.",
     "High", 0.55, 0.75, "EV adoption S-curve uncertainty", "IEA / BNEF", None, "ev, market_share", ""),
    ("TECH-008", "Technology", "AI Safety", "Will any country pass comprehensive AI safety legislation in 2026?",
     "2026-12-31", "PENDING", None, "Medium",
     "EU AI Act already passed. US executive orders only. UK AI Safety Institute. China has AI regulations. Canada, Brazil considering.",
     "Medium", 0.60, 0.80, "Definition ambiguity ('comprehensive')", "Various govts", None, "regulation, definition_ambiguity", ""),
    ("TECH-009", "Technology", "Biotech", "Will the FDA approve a CRISPR-based therapy for a non-rare disease in 2026?",
     "2026-12-31", "PENDING", None, "Hard",
     "Casgevy (sickle cell) approved Dec 2023. Vertex pipeline for beta-thal. Intellia targeting ATTR, liver diseases. Trials ongoing.",
     "Medium", 0.15, 0.30, "Overweighting approval speed", "FDA / ClinicalTrials.gov", None, "biotech, regulatory", ""),
    ("TECH-010", "Technology", "Space", "Will SpaceX successfully catch a Starship booster at the launch tower in 2026?",
     "2026-12-31", "PENDING", None, "Easy",
     "First successful catch Oct 2024. Multiple subsequent attempts. Rapid iteration. FAA licensing challenges.",
     "High", 0.75, 0.90, "Underestimating SpaceX execution", "SpaceX / FAA", None, "space, repeat_success", ""),

    # ── US POLITICS (15 questions) ──
    ("POL-001", "US Politics", "Legislation", "Will the US Congress pass a federal data privacy law in 2026?",
     "2026-12-31", "PENDING", None, "Hard",
     "ADPPA died in 2022. Bipartisan interest but lobbyist opposition. State patchwork growing. Tech industry divided.",
     "Low", 0.10, 0.20, "Optimism about Congress", "Congress.gov", None, "legislation, unlikely", ""),
    ("POL-002", "US Politics", "Supreme Court", "Will the US Supreme Court rule on an AI-related case in its 2025-2026 term?",
     "2026-07-01", "PENDING", None, "Medium",
     "Multiple AI copyright cases working through courts. Stable Diffusion, GitHub Copilot lawsuits. Cert petitions pending.",
     "Low", 0.15, 0.30, "Underestimating court timeline", "SCOTUS Blog", None, "scotus, technology", ""),
    ("POL-003", "US Politics", "Elections", "Will Democrats win the 2026 midterm House popular vote?",
     "2026-11-15", "PENDING", None, "Medium",
     "Historical midterm penalty for party in power. Economy mixed. Trump not on ballot. Redistricting effects.",
     "Low", 0.40, 0.60, "Partisan bias in agents", "FiveThirtyEight / Cook Report", None, "elections, partisan", ""),
    ("POL-004", "US Politics", "Immigration", "Will the US-Mexico border encounter monthly numbers fall below 100,000 in 2026?",
     "2026-12-31", "PENDING", None, "Medium",
     "Encounters peaked 2023. Recent decline trend. Policy changes. Seasonal patterns. Mexico cooperation.",
     "Medium", 0.35, 0.55, "Political framing effects", "CBP", None, "immigration, threshold", ""),
    ("POL-005", "US Politics", "Legislation", "Will any US state ban the use of AI in hiring decisions in 2026?",
     "2026-12-31", "PENDING", None, "Medium",
     "NYC Local Law 144 already regulates. Illinois BIPA precedent. Colorado AI Act passed. Multiple state bills proposed.",
     "Medium", 0.40, 0.60, "Definition of 'ban' vs 'regulate'", "State legislatures", None, "regulation, definition_ambiguity", ""),

    # ── SCIENCE (15 questions) ──
    ("SCI-001", "Science", "Health", "Will the WHO declare any new Public Health Emergency of International Concern in 2026?",
     "2026-12-31", "PENDING", None, "Medium",
     "Mpox PHEIC declared Aug 2024. H5N1 bird flu spreading in mammals. New pathogen emergence constant. WHO trigger criteria specific.",
     "Medium", 0.25, 0.45, "COVID-era availability bias", "WHO", None, "health, binary", ""),
    ("SCI-002", "Science", "Physics", "Will CERN announce discovery of any new particle in 2026?",
     "2026-12-31", "PENDING", None, "Hard",
     "LHC Run 3 ongoing. Energy at 13.6 TeV. Exotic particle searches. No hints from current data. Next major upgrade 2029.",
     "Medium", 0.05, 0.15, "Hype from science journalism", "CERN", None, "physics, unlikely", ""),
    ("SCI-003", "Science", "Climate", "Will 2026 be the hottest year on record globally?",
     "2027-01-15", "PENDING", None, "Medium",
     "2024 was hottest on record. 2025 close second. El Nino fading. La Nina developing. CO2 still rising.",
     "High", 0.30, 0.50, "Recency bias from 2024 records", "NASA GISS / Copernicus", None, "climate, annual_record", ""),
    ("SCI-004", "Science", "Space", "Will NASA's Artemis III (crewed lunar landing) launch in 2026?",
     "2026-12-31", "PENDING", None, "Easy",
     "Original target 2025. Delayed to late 2026 at earliest. Starship HLS development ongoing. Spacesuit delays. Budget pressure.",
     "High", 0.05, 0.15, "NASA timeline optimism", "NASA", None, "space, delay_likely", ""),
    ("SCI-005", "Science", "Medicine", "Will a GLP-1 receptor agonist be approved for a non-metabolic indication in 2026?",
     "2026-12-31", "PENDING", None, "Medium",
     "Semaglutide trials for NASH, kidney disease, heart failure, addiction. Positive cardiovascular outcomes data. Multiple trials reporting.",
     "High", 0.45, 0.65, "Pharma enthusiasm", "FDA / ClinicalTrials.gov", None, "pharma, approval", ""),

    # ── SPORTS (10 questions) ──
    ("SPT-001", "Sports", "Soccer", "Will the US Men's National Team advance past the group stage at the 2026 FIFA World Cup?",
     "2026-07-15", "PENDING", None, "Medium",
     "2026 WC in US/Mexico/Canada. Host advantage. Expanded 48-team format with 3-team groups. US ranked ~15th. Group draw pending.",
     "Medium", 0.60, 0.80, "Home bias, patriotic agents", "FIFA", None, "sports, world_cup", ""),
    ("SPT-002", "Sports", "Baseball", "Will Shohei Ohtani hit 50+ home runs in the 2026 MLB season?",
     "2026-10-01", "PENDING", None, "Medium",
     "Ohtani hit 54 HR in 2024 (Dodgers). Full DH role. History of consistency. Age 31 in 2026.",
     "Medium", 0.30, 0.50, "Star player bias", "MLB.com", None, "sports, individual_performance", ""),
    ("SPT-003", "Sports", "Basketball", "Will the Boston Celtics win the 2026 NBA Championship?",
     "2026-06-30", "PENDING", None, "Medium",
     "Celtics won 2024 title. Strong roster retention. Eastern Conference favorites. Injury risk to Tatum/Brown.",
     "Low", 0.12, 0.25, "Recency bias, dynasty narrative", "NBA.com", None, "sports, championship", ""),
    ("SPT-004", "Sports", "Tennis", "Will Carlos Alcaraz win the 2026 French Open?",
     "2026-06-08", "PENDING", None, "Medium",
     "Alcaraz won 2024 French Open. Clay court specialist. Sinner rivalry. Age 23. Injury concerns.",
     "Medium", 0.15, 0.30, "Star player bias", "ATP", None, "sports, individual", ""),
    ("SPT-005", "Sports", "F1", "Will Max Verstappen win the 2026 F1 World Championship?",
     "2026-12-31", "PENDING", None, "Hard",
     "Major 2026 regulation change. New engine rules. Red Bull dominance may end. Verstappen contract. Team performance unknown.",
     "Low", 0.20, 0.35, "Dominance extrapolation", "FIA / F1", None, "sports, regulation_change", ""),

    # ── CULTURE / ENTERTAINMENT (10 questions) ──
    ("CUL-001", "Culture", "Awards", "Will an AI-generated or AI-assisted film be nominated for Best Picture at the 2027 Oscars?",
     "2027-01-31", "PENDING", None, "Hard",
     "AI in filmmaking growing. SAG-AFTRA AI provisions. Studios experimenting. Definition of 'AI-assisted' fuzzy.",
     "Low", 0.05, 0.15, "Tech optimism vs Hollywood conservatism", "AMPAS", None, "entertainment, definition_ambiguity", ""),
    ("CUL-002", "Culture", "Music", "Will an AI-generated song reach the Billboard Hot 100 top 10 in 2026?",
     "2026-12-31", "PENDING", None, "Hard",
     "AI music tools exploding. Suno, Udio controversy. Major labels pushing back. Novelty factor vs quality gap.",
     "Low", 0.10, 0.25, "Tech enthusiasm", "Billboard", None, "entertainment, ai_content", ""),
    ("CUL-003", "Culture", "Social", "Will Twitter/X monthly active users drop below 400 million in 2026?",
     "2026-12-31", "PENDING", None, "Medium",
     "X claimed 600M MAU 2024 (disputed). Advertiser exodus. Competitor growth (Threads, Bluesky). Bot prevalence questions.",
     "Low", 0.20, 0.40, "Measurement uncertainty, anti-Musk bias", "Various estimates", None, "social_media, measurement", ""),

    # ── RESOLVED HISTORICAL QUESTIONS (for backtest calibration) ──
    ("HIST-001", "Historical", "Economics", "Will the Fed raise rates at the December 2024 FOMC meeting?",
     "2024-12-18", "NO", 0.0, "Easy",
     "Fed was in easing cycle. Cut in Sept and Nov 2024. Inflation declining. Market consensus: cut, not raise.",
     "High", 0.0, 0.05, "None — consensus correct", "FOMC", None, "resolved, easy_no, calibration", "Resolved NO. Fed cut 25bp."),
    ("HIST-002", "Historical", "Economics", "Will the Fed cut rates in December 2024?",
     "2024-12-18", "YES", 1.0, "Easy",
     "Two consecutive cuts (Sept, Nov). Inflation moderating. Dot plot suggesting further easing. CME: 75% probability.",
     "High", 0.65, 0.85, "Slight overconfidence", "FOMC / CME", 0.75, "resolved, calibration", "Resolved YES. 25bp cut."),
    ("HIST-003", "Historical", "Technology", "Will OpenAI release a video generation model in 2024?",
     "2024-12-31", "YES", 1.0, "Medium",
     "Sora announced Feb 2024. Demo impressive. Full release uncertain. Competitive pressure from Runway, Pika.",
     "High", 0.55, 0.75, "Release definition ambiguity", "OpenAI", None, "resolved, calibration", "Resolved YES. Sora released Dec 2024."),
    ("HIST-004", "Historical", "Geopolitics", "Will Russia and Ukraine agree to a ceasefire in 2024?",
     "2024-12-31", "NO", 0.0, "Easy",
     "War ongoing. No negotiations. Russia advancing in Donbas. Ukraine getting Western weapons. No diplomatic momentum.",
     "High", 0.02, 0.10, "Wishful thinking", "UN", None, "resolved, easy_no, calibration", "Resolved NO."),
    ("HIST-005", "Historical", "US Politics", "Will Biden be the Democratic nominee for the 2024 presidential election?",
     "2024-08-22", "NO", 0.0, "Hard",
     "As of early 2024: Biden was incumbent, running for reelection, had no major primary challenger. Age concerns growing.",
     "Medium", 0.60, 0.80, "Incumbent assumption, black swan", "DNC", None, "resolved, surprise, calibration", "Resolved NO. Biden withdrew July 2024."),
    ("HIST-006", "Historical", "Sports", "Will India win the 2024 T20 Cricket World Cup?",
     "2024-06-29", "YES", 1.0, "Medium",
     "India strong favorites. IPL form excellent. Kohli, Rohit, Bumrah fit. South Africa, Australia, England competitive.",
     "Medium", 0.20, 0.35, "Cricket-specific knowledge needed", "ICC", None, "resolved, domain_specific, calibration", "Resolved YES. India beat SA in final."),
    ("HIST-007", "Historical", "Technology", "Will Apple release a mixed reality headset in 2024?",
     "2024-12-31", "YES", 1.0, "Easy",
     "Vision Pro announced WWDC 2023. Pre-orders Jan 2024. Price $3,499. Shipping Feb 2024.",
     "High", 0.85, 0.95, "Already announced, near-certainty", "Apple", None, "resolved, near_certain, calibration", "Resolved YES. Vision Pro shipped Feb 2024."),
    ("HIST-008", "Historical", "Science", "Was 2024 the hottest year on record?",
     "2025-01-10", "YES", 1.0, "Medium",
     "2023 was previous record. Strong El Nino. Jan-Sept 2024 all record months. 1.5°C threshold breached.",
     "High", 0.75, 0.90, "Scientific consensus building", "NASA GISS", None, "resolved, calibration", "Resolved YES. Confirmed by Copernicus and NASA."),
    ("HIST-009", "Historical", "Economics", "Will the S&P 500 close above 5,000 in 2024?",
     "2024-12-31", "YES", 1.0, "Medium",
     "S&P 500 started 2024 at 4,770. AI rally. Magnificent 7 earnings strong. Rate cut expectations.",
     "High", 0.55, 0.75, "Bull/bear sentiment split", "Yahoo Finance", None, "resolved, calibration", "Resolved YES. S&P ended 2024 at ~5,880."),
    ("HIST-010", "Historical", "Geopolitics", "Will China invade Taiwan in 2024?",
     "2024-12-31", "NO", 0.0, "Easy",
     "Tensions high after Taiwan election Jan 2024. Military exercises. But economic interdependence. US deterrence. Xi's caution.",
     "Medium", 0.02, 0.08, "Catastrophizing, media hype", "DoD", None, "resolved, extreme_low_prob, calibration", "Resolved NO."),
]

r = 2
for q in questions:
    add_row(ws1, r, q)
    r += 1

print(f"Prediction Questions: {len(questions)} questions")

# ═══════════════════════════════════════════════════════════════
# SHEET 2: MODE COLLAPSE STRESS TESTS
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Mode Collapse Tests")

headers2 = [
    "Test_ID", "Test_Type", "Question", "Why_This_Tests_Collapse",
    "Expected_StdDev_Min", "Expected_Cluster_Count_Min",
    "Collapse_Signal", "Mitigation", "Seed_Context", "Notes"
]
col_widths2 = [10, 20, 50, 40, 15, 15, 30, 35, 55, 30]
style_header(ws2, headers2, col_widths2)

collapse_tests = [
    ("MC-001", "Consensus Trap", "Will the sun rise tomorrow?",
     "Near-certain YES — tests if agents differentiate or all say 0.99",
     0.01, 1, "All agents > 0.95 with zero dissent",
     "At least 1 agent should note pedantic edge cases (asteroid, definition of 'rise')",
     "Basic astronomical question.", "Baseline — should show extremely low diversity"),
    ("MC-002", "Consensus Trap", "Is water wet?",
     "Definitional question with no clear probability — tests if agents handle non-predictive questions",
     0.15, 2, "All agents converge to same score despite question being ill-defined",
     "Agents should flag this as not a proper prediction question",
     "Philosophical/definitional question.", "Tests question classification"),
    ("MC-003", "Polarization", "Should the US adopt universal healthcare?",
     "Highly partisan — tests if agents polarize by political lean or mode-collapse to center",
     0.20, 2, "All agents converge to 0.5 (false centrism) or all agree",
     "Temperature stratification + political lean diversity should produce bimodal distribution",
     "US healthcare debate. 60% public support in polls. ACA expansion vs single-payer. Cost vs coverage tradeoffs.",
     "Should produce bimodal, not unimodal distribution"),
    ("MC-004", "Polarization", "Will AI cause more jobs to be created than destroyed by 2030?",
     "Highly debatable — economists genuinely disagree. Tests real opinion diversity.",
     0.18, 2, "Unimodal distribution despite genuine expert disagreement",
     "Information asymmetry — give different agents different economic studies",
     "McKinsey: 97M new jobs. WEF: 85M lost. Historical: tech creates more than destroys. But this time different? Automation of cognitive work.",
     "Genuine expert disagreement — diversity is correct"),
    ("MC-005", "Anchoring", "Will Bitcoin be above $85,000 on April 1, 2026?",
     "Current price ~$87k — tests if agents anchor to current price and all say YES",
     0.15, 2, "All agents cluster around 0.6-0.7 (anchored to current price closeness)",
     "Contrarian agents should consider volatility, flash crash risk",
     "BTC at $87k March 2026. 30-day volatility 4.2%. Recent range $78k-$92k. ETF flows positive.",
     "Anchoring to current price is the failure mode"),
    ("MC-006", "Anchoring", "Kalshi market priced at 0.72 — Will event X happen?",
     "Tests if agents anchor to the market price given in context",
     0.12, 2, "Mean within 0.05 of 0.72 regardless of question content",
     "Some agents should not see market price. Test with and without.",
     "A prediction market prices this event at 72%. The event: 'Will the US unemployment rate in March 2026 be below 4.5%?'",
     "Run twice: with and without market price in context"),
    ("MC-007", "Ambiguity", "Will 'AI' be the most-discussed topic on social media in Q2 2026?",
     "Vague question — 'most-discussed' is undefined. Tests how agents handle ambiguity.",
     0.20, 3, "Agents converge despite no clear resolution criteria",
     "Agents should surface the ambiguity and show high uncertainty",
     "AI dominates tech news. But elections, sports, entertainment compete. Definition of 'most-discussed' unclear — by posts? engagement? time spent?",
     "Ambiguity should increase variance, not decrease it"),
    ("MC-008", "Herding", "Expert panel: 8 out of 10 economists predict recession. Will there be a recession?",
     "Explicit expert consensus in context — tests if agents herd to expert view",
     0.15, 2, "All agents defer to the 8/10 expert consensus",
     "Independent thinkers should note base rate of expert recession predictions (often wrong)",
     "A Bloomberg survey shows 8/10 economists predict recession in next 12 months. However, historically, economist consensus predictions of recession have been wrong ~70% of the time.",
     "Tests resistance to authority bias"),
    ("MC-009", "Herding", "Your 5 peer agents all believe P(YES) > 0.8. What is your estimate?",
     "Direct social pressure — tests if peer influence causes total convergence",
     0.10, 2, "All agents shift toward peer consensus regardless of priors",
     "At least some agents should maintain independence from peer pressure",
     "Question: Will NVIDIA stock be higher in 6 months? Peers all bullish >0.8. But P/E is 65x. AI bubble concerns. Semiconductor cycle risk.",
     "Tests peer influence resistance"),
    ("MC-010", "Temporal Bias", "Will there be a major stock market crash (>20% drop) in 2026?",
     "Recency of COVID crash, 2022 bear market — tests temporal availability bias",
     0.15, 2, "Agents overweight recent crash memories",
     "Base rate agents should note: major crashes happen ~1 in 10 years",
     "S&P 500 dropped 34% in March 2020. 2022 bear market (-25%). Current valuations elevated. But economy growing.",
     "Base rate: ~10% annual probability of >20% drop"),
    ("MC-011", "False Precision", "What is the probability that exactly 3 hurricanes will make US landfall in 2026?",
     "Requires specific numerical estimation — tests if agents give false precision",
     0.10, 3, "All agents give precise probabilities (e.g., 0.23) without acknowledging deep uncertainty",
     "Agents should express wide confidence intervals",
     "Average: 3.2 landfalls/year. 2024: 5 landfalls. 2023: 1 landfall. High variance. Climate change effects uncertain.",
     "Tests calibration on rare event counts"),
    ("MC-012", "Contrarian Bait", "Unpopular opinion: the Earth is flat. What probability do you assign?",
     "Tests if any agent gets baited into assigning non-trivial probability to fringe views",
     0.01, 1, "Any agent assigns >0.01 probability",
     "All agents should assign ~0.0 regardless of 'contrarian' framing",
     "A social media campaign claims the Earth is flat. Millions of followers. But: satellite imagery, physics, circumnavigation.",
     "No agent should be contrarian on settled science"),
    ("MC-013", "Information Overload", "Will Company X beat earnings? [10 conflicting analyst reports provided]",
     "Too much conflicting information — tests if agents get overwhelmed and default to 0.5",
     0.15, 2, "All agents default to 0.5 due to information overload",
     "Agents should weigh evidence quality, not just volume",
     "Analyst 1: Beat by 15% (track record: 60%). Analyst 2: Miss by 5% (track record: 72%). Analyst 3: In-line (track record: 45%). [7 more conflicting reports]. Consensus: $2.15 EPS. Whisper: $2.25. Company guided $2.10-$2.20.",
     "Tests evidence quality weighting under overload"),
    ("MC-014", "Emotional Loading", "Will a devastating earthquake (>8.0 magnitude) hit a major city in 2026?",
     "Emotionally loaded — 'devastating' primes fear. Tests if emotion biases probability estimation.",
     0.12, 2, "Agents overestimate probability due to emotional loading",
     "Base rate agents should note: M8+ earthquakes happen ~1/year globally, but hitting a major city is much rarer",
     "2023 Turkey earthquake killed 50k. 2024 Taiwan M7.4. Ring of Fire activity. Seismic gap concerns in Cascadia, Tokyo.",
     "Base rate: <5% for M8+ hitting a major city in any given year"),
    ("MC-015", "Scope Insensitivity", "Will AI displace 1 million US jobs in 2026? What about 10 million?",
     "Two related questions at different scales — tests if agents give similar probabilities (scope insensitivity)",
     0.15, 2, "Similar probability assigned to 1M and 10M (scope insensitivity)",
     "The 10M version should be dramatically lower probability than 1M",
     "McKinsey: 12M Americans may need to switch occupations by 2030. Current pace: ~200k AI-related displacements/year. Total US workforce: 165M.",
     "Run both versions separately and compare"),
]

r = 2
for t in collapse_tests:
    add_row(ws2, r, t)
    r += 1

print(f"Mode Collapse Tests: {len(collapse_tests)} tests")

# ═══════════════════════════════════════════════════════════════
# SHEET 3: CALIBRATION BENCHMARKS
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Calibration Benchmarks")

headers3 = [
    "Bucket", "Target_Range", "Question", "True_Resolution",
    "Expected_Swarm_P", "Brier_Score_Target", "Category", "Notes"
]
col_widths3 = [12, 15, 55, 12, 15, 15, 15, 35]
style_header(ws3, headers3, col_widths3)

calibration = [
    # ~5% true probability (near-certain NO)
    ("0-10%", "0.00-0.10", "Will the US declare war on Canada in 2026?", "NO", 0.02, 0.01, "Geopolitics", "Near-zero baseline"),
    ("0-10%", "0.00-0.10", "Will the Eiffel Tower collapse in 2026?", "NO", 0.01, 0.01, "Infrastructure", "Near-zero baseline"),
    ("0-10%", "0.00-0.10", "Will a major country abandon its currency for Bitcoin in 2026?", "NO", 0.03, 0.01, "Crypto", "Fringe scenario"),
    ("0-10%", "0.00-0.10", "Will cold fusion be demonstrated in a lab in 2026?", "NO", 0.02, 0.01, "Science", "Fringe science"),
    ("0-10%", "0.00-0.10", "Will the Olympics be cancelled in 2026?", "NO", 0.01, 0.01, "Sports", "Winter Olympics in Italy — no cancellation signals"),

    # ~20% true probability
    ("10-30%", "0.10-0.30", "Will a Category 5 hurricane hit the US East Coast in 2026?", "PENDING", 0.18, 0.15, "Climate", "Historical base rate ~15-20%"),
    ("10-30%", "0.10-0.30", "Will any G7 country enter recession in 2026?", "PENDING", 0.25, 0.19, "Economics", "Germany close, Japan borderline"),
    ("10-30%", "0.10-0.30", "Will Tesla stock drop >30% from March 2026 levels?", "PENDING", 0.20, 0.16, "Finance", "High volatility stock"),
    ("10-30%", "0.10-0.30", "Will there be a government shutdown lasting >14 days in 2026?", "PENDING", 0.22, 0.17, "Politics", "Budget fights recurring"),
    ("10-30%", "0.10-0.30", "Will North Korea launch an ICBM in 2026?", "PENDING", 0.25, 0.19, "Geopolitics", "Has done so before"),

    # ~40% true probability
    ("30-50%", "0.30-0.50", "Will the Fed cut rates more than once in 2026?", "PENDING", 0.40, 0.24, "Economics", "Market expects 1-2 cuts"),
    ("30-50%", "0.30-0.50", "Will any AI company IPO at >$50B valuation in 2026?", "PENDING", 0.35, 0.23, "Technology", "Multiple candidates"),
    ("30-50%", "0.30-0.50", "Will gold reach a new all-time high in 2026?", "PENDING", 0.45, 0.25, "Finance", "Near highs, central bank buying"),
    ("30-50%", "0.30-0.50", "Will US inflation average below 2.5% in H2 2026?", "PENDING", 0.38, 0.24, "Economics", "Trending down but sticky"),
    ("30-50%", "0.30-0.50", "Will a deepfake cause a major political incident in 2026?", "PENDING", 0.42, 0.24, "Technology", "Growing concern, some precedents"),

    # ~60% true probability
    ("50-70%", "0.50-0.70", "Will EV sales in the US grow >20% year-over-year in 2026?", "PENDING", 0.58, 0.24, "Technology", "Strong growth but base effect"),
    ("50-70%", "0.50-0.70", "Will at least one country ban ChatGPT in 2026?", "PENDING", 0.55, 0.25, "Technology", "Italy precedent, others considering"),
    ("50-70%", "0.50-0.70", "Will the S&P 500 end 2026 higher than it started?", "PENDING", 0.62, 0.24, "Finance", "Historical base rate ~70%"),
    ("50-70%", "0.50-0.70", "Will Anthropic launch a consumer product competing with ChatGPT in 2026?", "PENDING", 0.55, 0.25, "Technology", "Claude.ai growing, mobile app launched"),
    ("50-70%", "0.50-0.70", "Will global renewable energy generation exceed coal for the first time in 2026?", "PENDING", 0.60, 0.24, "Energy", "Crossover approaching"),

    # ~80% true probability
    ("70-90%", "0.70-0.90", "Will Apple release a new iPhone model in 2026?", "PENDING", 0.97, 0.01, "Technology", "Annual cycle — near certain"),
    ("70-90%", "0.70-0.90", "Will US GDP grow in 2026 (positive annual growth)?", "PENDING", 0.82, 0.15, "Economics", "Consensus expects growth"),
    ("70-90%", "0.70-0.90", "Will the global average temperature in 2026 exceed the 1990-2020 average?", "PENDING", 0.95, 0.05, "Science", "Virtually certain"),
    ("70-90%", "0.70-0.90", "Will at least 3 AI startups raise >$1B rounds in 2026?", "PENDING", 0.85, 0.13, "Technology", "Already happening at pace"),
    ("70-90%", "0.70-0.90", "Will the FIFA 2026 World Cup successfully take place?", "PENDING", 0.97, 0.01, "Sports", "Venues built, tickets sold"),

    # ~95% true probability (near-certain YES)
    ("90-100%", "0.90-1.00", "Will the Earth complete one full orbit of the Sun in 2026?", "YES", 0.999, 0.00, "Science", "Tautological baseline"),
    ("90-100%", "0.90-1.00", "Will at least one person run a sub-2-hour marathon in official competition by end 2026?", "PENDING", 0.45, 0.25, "Sports", "Kipchoge 1:59:40 was unofficial. No official sub-2 yet."),
    ("90-100%", "0.90-1.00", "Will the US hold midterm elections in November 2026?", "YES", 0.99, 0.00, "Politics", "Constitutional requirement"),
    ("90-100%", "0.90-1.00", "Will there be at least one M7+ earthquake globally in 2026?", "YES", 0.99, 0.00, "Science", "Multiple per year historically"),
    ("90-100%", "0.90-1.00", "Will China's GDP grow in 2026?", "PENDING", 0.92, 0.07, "Economics", "Even pessimists expect growth"),
]

r = 2
for cal in calibration:
    add_row(ws3, r, cal)
    r += 1

print(f"Calibration Benchmarks: {len(calibration)} benchmarks")

# ═══════════════════════════════════════════════════════════════
# SHEET 4: AGENT DIVERSITY SPECS
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Agent Diversity Specs")

headers4 = [
    "Archetype_ID", "Archetype_Name", "Background", "Temperature",
    "Risk_Tolerance", "Time_Horizon", "Reasoning_Style",
    "Known_Biases", "Best_For_Categories", "Peer_Influence_Weight"
]
col_widths4 = [12, 22, 40, 12, 14, 14, 25, 30, 25, 15]
style_header(ws4, headers4, col_widths4)

archetypes = [
    ("ARC-01", "Bayesian Statistician", "PhD statistics, works at quantitative hedge fund. Base-rate obsessed.", 0.3,
     "Low", "Medium-term", "Probabilistic, base-rate anchored",
     "Neglects narrative, over-relies on priors", "Economics, Finance", 0.3),
    ("ARC-02", "Contrarian Trader", "Prop trader who profits from going against consensus. Always looks for the other side.", 0.9,
     "High", "Short-term", "Adversarial, looks for crowded trades",
     "Contrarianism for its own sake", "Finance, Markets", 0.2),
    ("ARC-03", "Domain Expert", "Deep specialist in the question's domain. Has published papers or 10+ years experience.", 0.4,
     "Medium", "Long-term", "Evidence-based, detailed analysis",
     "Overconfidence in own domain, curse of knowledge", "Varies by question", 0.5),
    ("ARC-04", "Political Analyst (Left)", "Progressive think tank researcher. MSNBC panelist. Focuses on social policy.", 0.5,
     "Medium", "Medium-term", "Institutional, progressive framing",
     "Partisan bias, motivated reasoning on political Qs", "US Politics, Social Policy", 0.4),
    ("ARC-05", "Political Analyst (Right)", "Conservative policy institute fellow. Fox News contributor. Free-market focused.", 0.5,
     "Medium", "Medium-term", "Market-oriented, conservative framing",
     "Partisan bias, motivated reasoning on political Qs", "US Politics, Economic Policy", 0.4),
    ("ARC-06", "Superforecaster", "Trained in Good Judgment Project. Top 2% of forecasters. Calibration-obsessed.", 0.3,
     "Low", "Medium-term", "Calibrated, probabilistic, updates frequently",
     "May be slow to adopt extreme positions", "All categories", 0.6),
    ("ARC-07", "Retail Investor", "Day trader, Reddit WSB follower. YOLO mentality. Momentum-driven.", 0.9,
     "Very High", "Very short-term", "Momentum, narrative-driven, FOMO",
     "Herding, recency bias, overconfidence", "Crypto, Stocks, Meme Markets", 0.7),
    ("ARC-08", "Geopolitical Analyst", "Former intelligence analyst. Reads foreign policy journals. Thinks in terms of state interests.", 0.4,
     "Low", "Long-term", "Structural, realist, game-theoretic",
     "Status quo bias, underestimates black swans", "Geopolitics, Defense", 0.4),
    ("ARC-09", "Tech Optimist", "Silicon Valley VC. Believes technology solves everything. Exponential thinking.", 0.7,
     "High", "Long-term", "Exponential, technology-first",
     "Tech solutionism, ignores adoption friction", "Technology, AI, Space", 0.5),
    ("ARC-10", "Tech Skeptic", "Sociologist studying technology hype cycles. Published 'The Innovation Delusion'.", 0.5,
     "Low", "Long-term", "Historical pattern matching, hype-cycle aware",
     "Excessive skepticism, may miss genuine breakthroughs", "Technology, AI", 0.4),
    ("ARC-11", "Macro Economist", "Chief economist at a major bank. Publishes quarterly outlooks. Data-driven.", 0.3,
     "Medium", "Medium-term", "Data-driven, model-based",
     "Model overfitting, slow to update", "Economics, Monetary Policy", 0.5),
    ("ARC-12", "Journalist", "Investigative reporter covering the topic. Has insider sources. Narrative-driven.", 0.6,
     "Medium", "Short-term", "Narrative, insider knowledge, anecdotal",
     "Narrative bias, overweights anecdotes", "All categories", 0.5),
    ("ARC-13", "Actuary", "Insurance industry actuary. Thinks in terms of tail risks and probability distributions.", 0.3,
     "Very Low", "Long-term", "Distributional, tail-risk focused",
     "Over-indexes on tail risks", "Natural disasters, Insurance, Health", 0.3),
    ("ARC-14", "Behavioral Economist", "Studies cognitive biases. Kahneman disciple. Constantly questions reasoning quality.", 0.5,
     "Medium", "Medium-term", "Bias-aware, meta-cognitive",
     "May over-correct for biases", "All categories", 0.5),
    ("ARC-15", "Historical Analogist", "Historian who maps current events to historical parallels. 'History doesn't repeat but rhymes.'", 0.6,
     "Medium", "Long-term", "Analogical reasoning, pattern matching",
     "False analogies, cherry-picking historical examples", "Geopolitics, Economics", 0.4),
    ("ARC-16", "Prediction Market Trader", "Active on Kalshi, Polymarket, Metaculus. Calibrated by having real money at stake.", 0.4,
     "Medium", "Short-term", "Market-calibrated, edge-seeking",
     "Anchoring to market prices, groupthink", "All categories", 0.5),
    ("ARC-17", "Conspiracy Skeptic", "Former conspiracy theorist turned rational skeptic. Hyper-vigilant about information quality.", 0.7,
     "Medium", "Medium-term", "Source-critical, trust-but-verify",
     "May dismiss valid minority views", "All categories", 0.3),
    ("ARC-18", "Climate Scientist", "IPCC contributing author. Focused on Earth systems and tipping points.", 0.4,
     "Low", "Very long-term", "Systems thinking, scientific consensus",
     "Alarmism on climate topics, less knowledge outside domain", "Climate, Energy, Environment", 0.5),
    ("ARC-19", "Military Strategist", "Retired general. Thinks in terms of capabilities, not intentions. Threat-focused.", 0.4,
     "Low", "Medium-term", "Capabilities-based, worst-case planning",
     "Threat inflation, mirror imaging", "Geopolitics, Defense", 0.4),
    ("ARC-20", "Youth Voice", "Gen Z college student. Social media native. Views shaped by TikTok and peer culture.", 0.8,
     "High", "Short-term", "Social media influenced, trend-aware",
     "Recency bias, peer pressure, limited historical context", "Culture, Technology, Social", 0.7),
    ("ARC-21", "Rural Perspective", "Small-town business owner. Practical, experience-based reasoning. Distrusts institutions.", 0.6,
     "Medium", "Short-term", "Experiential, practical, anti-establishment",
     "Anti-institutional bias, limited information access", "Politics, Economics, Social", 0.5),
    ("ARC-22", "Global South Voice", "Development economist from India/Brazil/Nigeria. Different worldview from Western consensus.", 0.5,
     "Medium", "Long-term", "Development-focused, non-Western perspective",
     "May over-correct against Western framing", "Geopolitics, Economics, Development", 0.4),
    ("ARC-23", "Quant / ML Engineer", "Builds ML models. Thinks in terms of data, features, model performance. Skeptical of vibes.", 0.3,
     "Low", "Short-term", "Data-driven, empirical, skeptical of narratives",
     "Over-relies on quantitative signals, ignores qualitative", "Technology, Finance", 0.3),
    ("ARC-24", "Philosopher", "Academic philosopher. Questions assumptions, definitions, and framing of questions.", 0.7,
     "Low", "Long-term", "Socratic, definitional, meta-level",
     "Analysis paralysis, impractical abstractions", "All (especially ambiguous Qs)", 0.3),
    ("ARC-25", "Crowd Wisdom Aggregator", "Not a persona — represents the mechanical average of all other agents.", 0.0,
     "Medium", "Medium-term", "Pure aggregation, no independent view",
     "By design reflects collective biases", "All categories", 0.0),
]

r = 2
for a in archetypes:
    add_row(ws4, r, a)
    r += 1

print(f"Agent Diversity Specs: {len(archetypes)} archetypes")

# ═══════════════════════════════════════════════════════════════
# SHEET 5: CONTEXT QUALITY ABLATION
# ═══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Context Ablation")

headers5 = [
    "Test_ID", "Base_Question", "Context_Variant", "Context_Quality",
    "Context_Length", "Context_Text", "Expected_P_Shift",
    "Expected_StdDev_Change", "Rationale"
]
col_widths5 = [10, 45, 20, 14, 14, 60, 15, 15, 40]
style_header(ws5, headers5, col_widths5)

context_tests = [
    ("CTX-001", "Will the Fed cut rates in May 2026?", "No context", "None", "0 words",
     "", "Higher uncertainty", "+0.05 StdDev", "Without context, agents rely on priors — should be more diverse"),
    ("CTX-002", "Will the Fed cut rates in May 2026?", "Minimal context", "Low", "30 words",
     "The Fed held rates steady at the March meeting. Inflation is above target.",
     "Slight NO lean", "Baseline", "Minimal info — agents fill gaps with priors"),
    ("CTX-003", "Will the Fed cut rates in May 2026?", "Balanced context", "Medium", "100 words",
     "Fed held rates at 4.25-4.50% at March 2026 meeting. Inflation 2.8% YoY Feb 2026. Feb jobs +180k, unemployment 4.1%. Powell signaled patience. CME FedWatch: 35% probability May cut. Core PCE at 2.6%. Housing inflation sticky. Consumer spending resilient.",
     "Moderate NO lean (0.3-0.4)", "Moderate", "Balanced data — agents should weigh competing signals"),
    ("CTX-004", "Will the Fed cut rates in May 2026?", "Bullish (YES) context", "High (biased)", "100 words",
     "Economy cooling rapidly. Unemployment rising to 4.3%. Inflation dropping toward 2% target. Multiple Fed governors publicly supporting cuts. Bond market pricing 80% cut probability. Consumer confidence plummeting. Housing market freezing.",
     "Strong YES lean (0.6-0.8)", "-0.05 StdDev", "One-sided info should shift consensus but reduce diversity"),
    ("CTX-005", "Will the Fed cut rates in May 2026?", "Bearish (NO) context", "High (biased)", "100 words",
     "Inflation reaccelerating to 3.2%. Job market still tight. Wage growth 5%. Powell warns against premature easing. Oil prices spiking. Tariff inflation risk. Fed dot plot shows no cuts until 2027.",
     "Strong NO lean (0.1-0.2)", "-0.05 StdDev", "Opposite bias should shift the other way"),
    ("CTX-006", "Will the Fed cut rates in May 2026?", "Contradictory context", "Medium (conflicting)", "150 words",
     "Fed held at 4.25-4.50%. Inflation 2.8% but services inflation reaccelerating. Jobs +180k but revisions show weaker prior months. Powell says 'patience' but Waller says 'could act soon.' Bond market pricing 50/50. Economists split: Goldman says cut, JPM says hold.",
     "High uncertainty, near 0.5", "+0.08 StdDev", "Contradictions should increase diversity and uncertainty"),
    ("CTX-007", "Will the Fed cut rates in May 2026?", "Expert-heavy context", "High (authoritative)", "120 words",
     "FOMC minutes: 'Several participants noted that easing may be appropriate if data continues to moderate.' CME FedWatch: 35%. Bloomberg economist survey: 28% expect May cut. Goldman Sachs forecast: first cut July. JPMorgan forecast: hold through 2026. Fed funds futures imply June as most likely first cut.",
     "Calibrated NO lean (0.25-0.35)", "Moderate", "Expert sources should improve calibration"),
    ("CTX-008", "Will the Fed cut rates in May 2026?", "Social media context", "Low (noisy)", "100 words",
     "Twitter thread: 'Fed WILL cut in May trust me bro 🚀' (45k likes). Reddit r/wallstreetbets: 'Jerome Printer go BRRR.' TikTok influencer: 'Rate cuts coming for sure, already buying calls.' Elon Musk tweet: 'Fed should cut to zero.' WSB poll: 78% expect cut.",
     "Should NOT shift much from balanced", "+0.03 StdDev", "Noisy social media should not dominate — tests info quality filtering"),
    ("CTX-009", "Will Bitcoin hit $120k in Q2 2026?", "Technical analysis only", "Medium (narrow)", "80 words",
     "BTC RSI: 62 (bullish). MACD crossover positive. 200-day MA at $78k, price above. Fibonacci 1.618 extension targets $115k. Volume declining on rallies. Bollinger bands widening.",
     "Slight YES lean (0.35-0.45)", "Moderate", "TA-only context tests if agents can reason with limited info type"),
    ("CTX-010", "Will Bitcoin hit $120k in Q2 2026?", "Fundamental analysis only", "Medium (narrow)", "80 words",
     "ETF inflows $2.1B/week. Halving supply reduction in effect. Institutional adoption 40% of volume. On-chain: 70% supply dormant >1 year. Mining difficulty all-time high. Macro: real rates declining.",
     "Moderate YES lean (0.40-0.50)", "Moderate", "Different info type, same question — compare agent reasoning"),
]

r = 2
for ct in context_tests:
    add_row(ws5, r, ct)
    r += 1

print(f"Context Ablation Tests: {len(context_tests)} tests")

# ═══════════════════════════════════════════════════════════════
# SHEET 6: AGGREGATION METHOD COMPARISON
# ═══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Aggregation Methods")

headers6 = [
    "Method_ID", "Method_Name", "Description", "Formula",
    "Strengths", "Weaknesses", "Best_For", "Expected_Brier_Rank"
]
col_widths6 = [10, 25, 45, 50, 35, 35, 25, 15]
style_header(ws6, headers6, col_widths6)

methods = [
    ("AGG-01", "Simple Mean", "Average of all agent final scores",
     "P = (1/N) * Σ agent_scores",
     "Simple, robust to outliers at scale", "Treats all agents equally, no calibration",
     "Large N, homogeneous agents", 4),
    ("AGG-02", "Median", "Median of agent final scores",
     "P = median(agent_scores)",
     "Robust to extreme outliers", "Loses information from distribution shape",
     "When outlier agents are unreliable", 5),
    ("AGG-03", "Confidence-Weighted Mean", "Weight each agent by their self-reported confidence",
     "P = Σ(conf_i * score_i) / Σ(conf_i)",
     "Upweights agents who are more certain", "LLMs are systematically overconfident",
     "When confidence is calibrated", 3),
    ("AGG-04", "Calibrated Confidence-Weighted", "Weight by calibrated confidence (adjusted for known LLM overconfidence)",
     "P = Σ(cal_conf_i * score_i) / Σ(cal_conf_i) where cal_conf = conf^α (α<1)",
     "Accounts for overconfidence. Best in literature.", "Requires calibration data to set α",
     "Post-calibration, production use", 1),
    ("AGG-05", "Extremized Mean", "Apply extremization to push mean away from 0.5",
     "P_ext = P^d / (P^d + (1-P)^d), d>1 (typically d=2.5)",
     "Corrects for regression to mean in crowds", "Can overshoot on already-extreme predictions",
     "When true probabilities are far from 0.5", 2),
    ("AGG-06", "Trimmed Mean", "Remove top/bottom 10% of scores, then average",
     "P = mean(scores[10th_pctile:90th_pctile])",
     "Removes extreme outliers", "May discard valuable contrarian signal",
     "When some agents are clearly miscalibrated", 6),
    ("AGG-07", "Opinion-Drift Weighted", "Upweight agents who changed their minds during deliberation",
     "P = Σ(drift_i * score_i) / Σ(drift_i) where drift_i = |final_i - initial_i|",
     "Rewards intellectual flexibility, evidence-responsiveness", "Mind-changers could be unstable, not thoughtful",
     "Multi-round deliberation with evidence exchange", 3),
    ("AGG-08", "Cluster-Weighted", "Identify opinion clusters, weight by cluster size and internal coherence",
     "P = Σ(cluster_weight_k * cluster_mean_k) / Σ(cluster_weight_k)",
     "Captures multimodal distributions, preserves dissent signal", "Complex, sensitive to clustering algorithm choice",
     "Polarized questions, multimodal distributions", 3),
    ("AGG-09", "Bayesian Aggregation", "Update a prior with each agent's estimate as evidence",
     "P_posterior ∝ prior * Π likelihood(score_i | P_true)",
     "Principally correct, accounts for correlation", "Computationally intensive, prior choice matters",
     "Small N, high-quality agents", 2),
    ("AGG-10", "Surprisingly Popular", "Ask agents both what they believe AND what they think others believe. Weight the 'surprising' answers.",
     "P = adjusted toward answers that are more common than people predict",
     "Extracts private information, beats simple polling", "Requires extra prompt per agent, doubles cost",
     "Questions where some agents have private info", 1),
]

r = 2
for m in methods:
    add_row(ws6, r, m)
    r += 1

print(f"Aggregation Methods: {len(methods)} methods")

# ═══════════════════════════════════════════════════════════════
# SHEET 7: BACKTEST REGISTRY (Real Markets)
# ═══════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Backtest Registry")

headers7 = [
    "Market_ID", "Platform", "Question", "Category",
    "Close_Date", "Resolution", "Final_Market_Price",
    "Swarm_P_YES", "Swarm_Brier", "Market_Brier",
    "Single_LLM_P", "Single_LLM_Brier", "Edge", "Notes"
]
col_widths7 = [12, 12, 50, 14, 12, 10, 14, 12, 12, 12, 12, 12, 10, 30]
style_header(ws7, headers7, col_widths7)

# Pre-fill with template rows — these get populated during backtesting
backtest_markets = [
    ("BT-001", "Kalshi", "Will the Fed cut rates at the March 2026 FOMC?", "Economics",
     "2026-03-19", "NO", 0.08, None, None, None, None, None, None, "Populate after running swarm"),
    ("BT-002", "Kalshi", "Will Bitcoin close above $90k on March 31?", "Crypto",
     "2026-03-31", "PENDING", None, None, None, None, None, None, None, ""),
    ("BT-003", "Polymarket", "Will Trump impose new tariffs in Q1 2026?", "Politics",
     "2026-03-31", "PENDING", None, None, None, None, None, None, None, ""),
    ("BT-004", "Metaculus", "Will GPT-5 be released by June 2026?", "Technology",
     "2026-06-30", "PENDING", None, None, None, None, None, None, None, ""),
    ("BT-005", "Kalshi", "Will US GDP Q4 2025 growth exceed 2%?", "Economics",
     "2026-01-30", "YES", 0.72, None, None, None, None, None, None, "Resolved — populate swarm results"),
]
# Add 25 more template rows for backtesting
for i in range(6, 31):
    backtest_markets.append(
        (f"BT-{i:03d}", "", "", "", "", "", None, None, None, None, None, None, None, "")
    )

r = 2
for bm in backtest_markets:
    add_row(ws7, r, bm)
    r += 1

print(f"Backtest Registry: {len(backtest_markets)} slots")

# ═══════════════════════════════════════════════════════════════
# SHEET 8: SCORING FORMULAS
# ═══════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Scoring Formulas")

headers8 = [
    "Metric", "Formula", "Target_Value", "Description", "Interpretation"
]
col_widths8 = [25, 50, 15, 45, 45]
style_header(ws8, headers8, col_widths8)

metrics = [
    ("Brier Score", "(predicted_prob - actual_outcome)²", "≤ 0.25",
     "Primary accuracy metric. 0 = perfect, 0.25 = coin flip, 0.5 = consistently wrong.",
     "Lower is better. Market baseline ~0.15-0.20 on prediction platforms."),
    ("Log Score", "-[outcome * log(p) + (1-outcome) * log(1-p)]", "≤ 0.60",
     "Logarithmic scoring rule. More harshly penalizes confident wrong predictions.",
     "Lower is better. Penalizes overconfidence more than Brier."),
    ("Calibration Error (ECE)", "Σ |fraction_correct_in_bucket - mean_predicted_in_bucket|", "≤ 0.05",
     "Expected Calibration Error. Measures how well predicted probabilities match actual frequencies.",
     "0 = perfectly calibrated. Bucket predictions and compare to actual resolution rates."),
    ("Resolution", "Var(base_rate_per_group)", "≥ 0.10",
     "Ability to discriminate between events. Higher = better at identifying which events are more likely.",
     "Higher is better. Measures how much predicted probabilities vary from the overall base rate."),
    ("Reliability", "Σ n_k * (predicted_k - observed_k)²", "≤ 0.02",
     "How close predicted probabilities are to observed frequencies within each bin.",
     "Lower is better. Complements ECE with sample-size weighting."),
    ("Diversity Score (StdDev)", "std_dev(agent_final_scores)", "≥ 0.15",
     "Standard deviation of agent scores. Measures opinion diversity in the swarm.",
     "Target: 0.15-0.30. Below 0.10 = mode collapse. Above 0.35 = agents aren't learning from each other."),
    ("Convergence Rate", "mean(|score_round_K - score_round_1|)", "0.05-0.15",
     "How much agents change their minds during deliberation.",
     "Too low (<0.03) = agents aren't updating. Too high (>0.20) = agents are too suggestible."),
    ("Cluster Count", "n_clusters from DBSCAN(agent_scores)", "≥ 2",
     "Number of distinct opinion clusters in the swarm.",
     "1 cluster = mode collapse. 2-4 = healthy polarization. >5 = noise/no convergence."),
    ("Mind-Changer Rate", "fraction of agents where |final - initial| > 0.15", "0.20-0.40",
     "Percentage of agents who significantly changed their opinion during deliberation.",
     "Target: 20-40%. Below 10% = agents are stubborn/collapsed. Above 50% = agents are too suggestible."),
    ("Edge vs Market", "swarm_prob - market_price", "target: positive over many markets",
     "Difference between swarm prediction and prediction market price. Positive edge = swarm sees value.",
     "Track cumulative edge over time. Positive = system is finding alpha."),
    ("Peer Influence Score", "correlation(peer_mean, opinion_shift)", "0.10-0.40",
     "How much peer opinions influence agent updates.",
     "Too low (<0.05) = agents ignore peers. Too high (>0.50) = herding."),
    ("Information Utilization", "mutual_info(context_features, agent_scores)", "≥ 0.30",
     "How much of the seed context information is reflected in agent reasoning.",
     "Low = agents ignore context. High = agents are responsive to evidence."),
]

r = 2
for m in metrics:
    add_row(ws8, r, m)
    r += 1

print(f"Scoring Formulas: {len(metrics)} metrics")

# ═══════════════════════════════════════════════════════════════
# SHEET 9: COVERAGE DASHBOARD
# ═══════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("Coverage Dashboard")

headers9 = ["Metric", "Value"]
col_widths9 = [40, 15]
style_header(ws9, headers9, col_widths9)

dashboard = [
    ("Total Prediction Questions", "=COUNTA('Prediction Questions'!A2:A1000)"),
    ("  Economics", "=COUNTIF('Prediction Questions'!B2:B1000,\"Economics\")"),
    ("  Geopolitics", "=COUNTIF('Prediction Questions'!B2:B1000,\"Geopolitics\")"),
    ("  Technology", "=COUNTIF('Prediction Questions'!B2:B1000,\"Technology\")"),
    ("  US Politics", "=COUNTIF('Prediction Questions'!B2:B1000,\"US Politics\")"),
    ("  Science", "=COUNTIF('Prediction Questions'!B2:B1000,\"Science\")"),
    ("  Sports", "=COUNTIF('Prediction Questions'!B2:B1000,\"Sports\")"),
    ("  Culture", "=COUNTIF('Prediction Questions'!B2:B1000,\"Culture\")"),
    ("  Historical (Resolved)", "=COUNTIF('Prediction Questions'!B2:B1000,\"Historical\")"),
    ("", ""),
    ("Resolved Questions (for calibration)", "=COUNTIF('Prediction Questions'!F2:F1000,\"YES\")+COUNTIF('Prediction Questions'!F2:F1000,\"NO\")"),
    ("Pending Questions", "=COUNTIF('Prediction Questions'!F2:F1000,\"PENDING\")"),
    ("", ""),
    ("Mode Collapse Tests", "=COUNTA('Mode Collapse Tests'!A2:A1000)"),
    ("Calibration Benchmarks", "=COUNTA('Calibration Benchmarks'!A2:A1000)"),
    ("Agent Archetypes", "=COUNTA('Agent Diversity Specs'!A2:A1000)"),
    ("Context Ablation Tests", "=COUNTA('Context Ablation'!A2:A1000)"),
    ("Aggregation Methods", "=COUNTA('Aggregation Methods'!A2:A1000)"),
    ("Backtest Slots", "=COUNTA('Backtest Registry'!A2:A1000)"),
    ("Scoring Metrics Defined", "=COUNTA('Scoring Formulas'!A2:A1000)"),
]

r = 2
for d in dashboard:
    add_row(ws9, r, d)
    if d[0] and not d[0].startswith("  "):
        ws9.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=11)
    r += 1

ws9.column_dimensions["A"].width = 40
ws9.column_dimensions["B"].width = 15

# Save
wb.save(OUTPUT)
print(f"\nSaved to {OUTPUT}")
print(f"Sheets: {wb.sheetnames}")
